import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import text
from config import get_engine
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule

os.makedirs('charts', exist_ok=True)
os.makedirs('exports', exist_ok=True)

def run_query(query):
    """Execute SQL query and return DataFrame"""
    engine = get_engine()
    with engine.connect() as conn:
        df = pd.read_sql(text(query), conn)
    print(f"Rows retrieved: {len(df)}")
    return df

# 1. pie chart - Customer Distribution by Type
def pie_chart():
    query = """
    SELECT ct.type_name, COUNT(c.customer_id) AS count
    FROM customers c
    JOIN customer_types ct ON c.customer_type_id = ct.customer_type_id
    GROUP BY ct.type_name;
    """
    df = run_query(query)
    
    plt.figure(figsize=(10, 7))
    plt.pie(df['count'], labels=df['type_name'], autopct='%1.1f%%')
    plt.title('Customer Distribution by Type')
    plt.savefig('charts/01_pie_customer_types.png', bbox_inches='tight')
    plt.close()
    print("Created: Pie chart - Customer types distribution\n")

# 2. bar chart - Top Branches by Transactions
def bar_chart():
    query = """
    SELECT b.branch_name, COUNT(t.transaction_id) AS tx_count
    FROM transactions t
    JOIN branches b ON t.branch_id = b.branch_id
    GROUP BY b.branch_name
    ORDER BY tx_count DESC
    LIMIT 10;
    """
    df = run_query(query)
    
    plt.figure(figsize=(10, 6))
    plt.bar(df['branch_name'], df['tx_count'], color='steelblue')
    plt.xlabel('Branch Name')
    plt.ylabel('Transaction Count')
    plt.title('Top 10 Branches by Transaction Count')
    plt.xticks(rotation=45, ha='right')
    plt.savefig('charts/02_bar_branches.png', bbox_inches='tight')
    plt.close()
    print("Created: Bar chart - Top branches by activity\n")

# 3. horizontal bar - Avg Amount by Transaction Type
def horizontal_bar_chart():
    query = """
    SELECT tt.type_name, AVG(t.amount)::NUMERIC(18,2) AS avg_amount
    FROM transactions t
    JOIN transaction_types tt ON t.transaction_type_id = tt.transaction_type_id
    GROUP BY tt.type_name
    ORDER BY avg_amount DESC;
    """
    df = run_query(query)
    
    plt.figure(figsize=(10, 6))
    plt.barh(df['type_name'], df['avg_amount'], color='coral')
    plt.xlabel('Average Amount ($)')
    plt.ylabel('Transaction Type')
    plt.title('Average Transaction Amount by Type')
    plt.savefig('charts/03_horizontal_bar_tx_types.png', bbox_inches='tight')
    plt.close()
    print("Created: Horizontal bar - Avg transaction amounts\n")

# 4. line chart - Monthly Transaction Trends
def line_chart():
    query = """
    SELECT DATE_TRUNC('month', t.transaction_date)::DATE AS month,
           COUNT(t.transaction_id) AS tx_count
    FROM transactions t
    WHERE t.transaction_date >= CURRENT_DATE - INTERVAL '12 months'
    GROUP BY month
    ORDER BY month;
    """
    df = run_query(query)
    
    plt.figure(figsize=(12, 6))
    plt.plot(df['month'], df['tx_count'], marker='o', linewidth=2)
    plt.xlabel('Month')
    plt.ylabel('Transaction Count')
    plt.title('Monthly Transaction Trends')
    plt.xticks(rotation=45)
    plt.grid(True, alpha=0.3)
    plt.savefig('charts/04_line_monthly_trends.png', bbox_inches='tight')
    plt.close()
    print("Created: Line chart - Monthly transaction trends\n")

# 5. histogram - Account Balance Distribution
def histogram():
    query = """
    SELECT a.balance
    FROM accounts a
    JOIN account_statuses ast ON a.account_status_id = ast.account_status_id
    WHERE LOWER(ast.status_name) IN ('active', 'open')
    AND a.balance BETWEEN -10000 AND 50000;
    """
    df = run_query(query)
    
    plt.figure(figsize=(10, 6))
    plt.hist(df['balance'], bins=30, color='green', alpha=0.7)
    plt.xlabel('Account Balance ($)')
    plt.ylabel('Number of Accounts')
    plt.title('Account Balance Distribution')
    plt.savefig('charts/05_histogram_balances.png', bbox_inches='tight')
    plt.close()
    print("Created: Histogram - Balance distribution\n")

# 6. scatter plot - Customer Age vs Balance
def scatter_plot():
    query = """
    SELECT EXTRACT(YEAR FROM AGE(CURRENT_DATE, c.date_of_birth)) AS age,
           SUM(a.balance)::NUMERIC(18,2) AS total_balance
    FROM customers c
    JOIN accounts a ON c.customer_id = a.customer_id
    WHERE c.date_of_birth IS NOT NULL
    GROUP BY c.customer_id, c.date_of_birth
    HAVING SUM(a.balance) > 0;
    """
    df = run_query(query)
    
    plt.figure(figsize=(10, 6))
    plt.scatter(df['age'], df['total_balance'], alpha=0.5)
    plt.xlabel('Customer Age')
    plt.ylabel('Total Balance ($)')
    plt.title('Customer Age vs Total Balance')
    plt.grid(True, alpha=0.3)
    plt.savefig('charts/06_scatter_age_balance.png', bbox_inches='tight')
    plt.close()
    print("Created: Scatter plot - Age vs Balance\n")

# 7. time slider - Interactive Transaction Trends
def time_slider_chart():
    query = """
    SELECT DATE_TRUNC('month', t.transaction_date)::DATE AS month,
           b.branch_name,
           SUM(t.amount)::NUMERIC(18,2) AS total_amount,
           COUNT(t.transaction_id) AS tx_count
    FROM transactions t
    JOIN branches b ON t.branch_id = b.branch_id
    GROUP BY month, b.branch_name
    ORDER BY month, b.branch_name;
    """
    df = run_query(query)
    df['month'] = df['month'].astype(str)

    fig = px.bar(
        df,
        x="branch_name",
        y="total_amount",
        color="branch_name",
        animation_frame="month",
        title="Branch Transaction Amounts Over Time",
        labels={"branch_name": "Branch", "total_amount": "Total Amount ($)"},
        height=600
    )

    filepath = "charts/time_slider.html"
    fig.write_html(filepath, auto_open=False)
    print(f"Saved: Interactive time slider chart → {filepath}\n")


# 8. excel export
def export_to_excel(dataframes_dict, filename):
    filepath = f'exports/{filename}'
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        total_rows = 0
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            total_rows += len(df)
    
    wb = load_workbook(filepath)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for col_idx in range(1, ws.max_column + 1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            sample = ws.cell(row=2, column=col_idx).value
            
            if isinstance(sample, (int, float)):
                rule = ColorScaleRule(
                    start_type="min", start_color="F8696B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="63BE7B"
                )
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{ws.max_row}", rule)
    
    wb.save(filepath)
    print(f"Excel created: {filename}, {len(dataframes_dict)} sheets, {total_rows} rows\n")

def main():
    print("\n=== ANALYTICS START ===\n")
    
    # Generate 6 charts
    pie_chart()
    bar_chart()
    horizontal_bar_chart()
    line_chart()
    histogram()
    scatter_plot()

    time_slider_chart()
    
    engine = get_engine()
    
    query1 = """
    SELECT c.customer_id, c.first_name, c.last_name,
           ct.type_name AS customer_type,
           COUNT(a.account_id) AS accounts,
           SUM(a.balance)::NUMERIC(18,2) AS total_balance
    FROM customers c
    JOIN customer_types ct ON c.customer_type_id = ct.customer_type_id
    LEFT JOIN accounts a ON c.customer_id = a.customer_id
    GROUP BY c.customer_id, c.first_name, c.last_name, ct.type_name;
    """
    
    query2 = """
    SELECT at.type_name AS account_type,
           COUNT(a.account_id) AS count,
           AVG(a.balance)::NUMERIC(18,2) AS avg_balance
    FROM accounts a
    JOIN account_types at ON a.account_type_id = at.account_type_id
    GROUP BY at.type_name;
    """
    
    with engine.connect() as conn:
        df1 = pd.read_sql(text(query1), conn)
        df2 = pd.read_sql(text(query2), conn)
    
    export_to_excel({
        'Customers': df1,
        'Accounts': df2
    }, 'banking_report.xlsx')
    
    print("=== DONE ===\n")

if __name__ == "__main__":
    main()